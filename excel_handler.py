import openpyxl

def read_products(excel_path):
    """Read product names from Excel"""
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active
    
    products = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
        if row[0]:  # If product name exists
            products.append(row[0])
    
    wb.close()
    return products

def write_prices(excel_path, prices_to_write, price_mode='min'):
    """Write prices to Excel based on mode (min/max)"""
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active
    
    for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        product_name = row[0].value
        
        if product_name in prices_to_write:
            price_list = prices_to_write[product_name]
            
            if price_mode == 'min':
                price = min(price_list)
            else:  # max
                price = max(price_list)
            
            sheet.cell(row=idx, column=2).value = price  # Column B
    
    wb.save(excel_path)
    wb.close()

# Test
if __name__ == "__main__":
    products = read_products("sample.xlsx")
    print(products)
